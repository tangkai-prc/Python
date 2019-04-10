#coding: utf-8

import time 
import os
import re
from hashlib import md5, sha1
from zlib import crc32
from openpyxl import Workbook
from openpyxl.styles import colors, Font, Fill, NamedStyle 
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import load_workbook 
from openpyxl.writer.excel import ExcelWriter
import xml.etree.ElementTree as ET

def getfileCreateTime(file):
    return time.ctime(os.path.getmtime(file))

def searchFile(FileType):
    filelist=[]
    for root, dirs, files in os.walk(".", topdown=False):
        for name in files:
            str = os.path.normpath(os.path.join(root, name))
            if str.split('.')[-1] == FileType:
                filelist.append(str)

    if len(filelist) == 0:
        return None
    else:
        return filelist

def getICDVersion(fileName):
    icd = ET.parse(fileName)
    root = icd.getroot()
    version = None
    for child in root:
        # 第二层节点的标签名称和属性.
        #print(child.tag,":", child.attrib) 
        for item in child.attrib:
            if item == 'configVersion' :    #item是attribute返回的字段?
                version =  child.attrib[item]
    return version

def getMd5(filename): #计算md5
    mdfive = md5()
    with open(filename, 'rb') as f:
        mdfive.update(f.read())
    return str.upper(mdfive.hexdigest())

def getCrc32(filename): #计算crc32
    with open(filename, 'rb') as f:
        res = hex(crc32(f.read()))
    return str.upper(res).replace('0X','')

class JoiFileVersion(object):
    def __init__(self, filename):
        self.__filename = filename
        self.__version = "V0.00"
        self.__crc     = "FFFFFFFF"
        self.__subq    = "RP30000000"
        self.__date    = "2000-01-01 00:00:00"
        self.__getDevVersionInfo()

    def getFileName(self):
        print(self.__filename)

    def __getDevVersionInfo(self):
        filename = self.__filename
        info = {}
        if False == os.path.exists(filename) :
            print("config文件不存在!")
            return None
        with open(filename , 'rt') as f:
            line = " "
            data = " "
            ppcData = " "
            while line != '':
                line = f.readline()
                if line.startswith("[FILE VERSION"):
                    data = ''.join(re.findall(r'\[(.+?)\]', line))
                    continue
                
                if line.startswith("[PPC VERSION="):
                    ppcData = ''.join(re.findall(r'\[(.+?)\]', line))
                    continue

        data = data.split(" ")
        ppcData = ppcData.split(" ")

        info['version']  = data[1].split('=')[1]
        info['Subq']  = data[2].split('=')[1]
        info['date']  = data[3].split('=')[1] + " " + data[4].split('=')[1]
        info['crc']  = data[6].split('=')[1]
        info['ppcversion']  = ppcData[1].split('=')[1]
        info['ppcdate']  = ppcData[2].split('=')[1] + " " + ppcData[3].split('=')[1]
        info['ppccrc']  = ppcData[4].split('=')[1]
        self.__version = info['version'][:5]
        self.__subq    = info['Subq']
        self.__date    = info['date']
        self.__crc     = info['crc']
        self.__ppcVersion = info['ppcversion']
        self.__ppcDate    = info['ppcdate']
        self.__ppcCrc     = info['ppccrc']  

        return info

    def getVersion(self):
        return self.__version
    
    def getSubq(self):
        return self.__subq

    def getDate(self):
        return self.__date
    
    def getCrc(self):
        return self.__crc

    def getPPCVersion(self):
        return self.__ppcVersion

    def getPPCDate(self):
        return self.__ppcDate

    def getPPCCrc(self):
        return self.__ppcCrc    

#对整个表进行设置样式设计
def setStyle(sheet, rows, columns):
    sheet.insert_rows(rows)
    # 字体
    font = Font(name='宋体', size=12, b=False)

    # 边框
    line_t = Side(style='thin', color='000000')  # 细边框
    line_m = Side(style='thick', color='000000')  # 粗边框

    border = Border(top=line_m, bottom=line_m, left=line_m, right=line_m)

    # 填充,无
    fill = PatternFill('solid', fgColor='CFCFCF')

    # 对齐
    alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

    #打包样式
    sty = NamedStyle(name='sty', font=font, border=border, alignment=alignment) #, fill=fill)

    for r in range(3, rows+1):
        sheet.row_dimensions[rows].height = 45
        for c in range(1, columns):
            if rows < 3 :
                pass
            else:
                sheet.column_dimensions[get_column_letter(c)].width = 15
                try:
                    sheet.cell(r, c).style = sty 
                except ValueError:
                    sheet.cell(r, c).style = 'sty' #Once registered assign the style using just the name:
                                               #ws['D5'].style = 'highlight'
                

def setDevCategory(keyName):
    #初始化表头列�?
    dictCategory = {
        #   key      : #value
        u"NSR-3611"  : u"线路保护",
        u"NSR-3620"  : u"电容器保护",
        u"NSR-3670"  : u"电抗器保护",
        u"NSR-3697"  : u"所用变保护",
        u"NSR-3613"  : u"母线保护",
        u"NSR-3641"  : u"备自投",
        u"NSR-3641RF" : u"备自投",
        u"NS-3641"   : u"备自投",
        u"NSR-378LR" : u"变压器后备保护"
    }
    for key in dictCategory:
        if keyName == key:
            return dictCategory[key]
    return "NULL"


if __name__ == '__main__':
    print("Starting......")
    bookName = searchFile("xlsx" or 'xls')
    patten = re.compile(u'[A-Za-z0-9-]')
    manufactureName = ''.join(patten.findall(bookName[0].strip('.xlsx'or'.xls')))
    deviceName = re.sub(u'-A\d+', '', manufactureName) #去掉硬件编码
    seriesName = deviceName.split('-')[0] +'-'+deviceName.split('-')[1].strip('A')
    #取sheet
    workbook = load_workbook(bookName[0])
    sheetnames = workbook.get_sheet_names() #获得表单名字
    CurrentSheet = workbook.get_sheet_by_name(sheetnames[0])
    setStyle(CurrentSheet, 3, column_index_from_string('AN')+1) #列号转换为数字
    #取config处理的方法
    config = JoiFileVersion('config.txt')
    
    CurrentSheet['A3'] = setDevCategory(seriesName)
    CurrentSheet['B3'] = seriesName
    CurrentSheet['C3'] = manufactureName    #硬件型号
    CurrentSheet['D3'] = deviceName     #应用子型号
    CurrentSheet['E3'] = " \n".join(searchFile("joi"))
    CurrentSheet['F3'] = " \n".join(searchFile("icd"))
    icdVersion = [getICDVersion(files) for files in searchFile("icd")] #多个文件创建时间需要合并打开?
    CurrentSheet['G3'] = " \n".join(icdVersion)
    icdCrc32 = [getCrc32(files) for files in searchFile("icd")] #多个文件创建时间需要合并打开?
    CurrentSheet['H3'] = " /\n".join(icdCrc32)
    icdMd5 = [getMd5(files) for files in searchFile("icd")] #多个文件创建时间需要合并打开?
    CurrentSheet['I3'] = " /\n".join(icdMd5)
    icdCreateTime = [getfileCreateTime(files) for files in searchFile("icd")] #多个文件创建时间需要合并打开?
    CurrentSheet['J3'] = " /\n".join(icdCreateTime)
    CurrentSheet['K3'] = config.getVersion()
    CurrentSheet['L3'] = config.getDate()
    joiCreateTime = [getfileCreateTime(files) for files in searchFile("joi")] #多个文件创建时间需要合并打开?
    CurrentSheet['M3'] = " /\n".join(joiCreateTime)
    CurrentSheet['N3'] = config.getSubq()
    CurrentSheet['O3'] = config.getCrc()
    CurrentSheet['Q3'] = config.getPPCVersion()
    CurrentSheet['R3'] = config.getPPCDate()
    CurrentSheet['S3'] = config.getPPCCrc()

    try:    
        workbook.save(bookName[0]) 
        workbook.close()
    except Exception as err:
        print("文件保存异常 %s"%(err))

    print('Succeed!')
